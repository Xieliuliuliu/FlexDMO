import sys
import time
import os
from collections import defaultdict
from scipy.stats import ranksums
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from utils.metrics import calculate_MIGD
from utils.result_io import load_result_from_files

def extract_info_from_settings(settings):
    """从settings中提取算法名、问题名、n和tau
    
    Args:
        settings: 包含配置信息的字典
        
    Returns:
        tuple: (响应策略, 搜索算法, 问题名, n, tau)
    """
    response = settings.get('response_strategy_class', 'Unknown')
    search = settings.get('search_algorithm_class', 'Unknown')
    problem = settings.get('problem_class', 'Unknown')
    
    # 从problem_params中获取n和tau
    problem_params = settings.get('problem_params', {})
    n = problem_params.get('n', None)
    tau = problem_params.get('tau', None)
    
    return response, search, problem, n, tau

def find_first_and_second(mean_list):
    """找出列表中的最小值和次小值
    
    Args:
        mean_list: 均值列表
        
    Returns:
        min1_index: 最小值的索引
        min2_index: 次小值的索引
    """
    if mean_list[0] < mean_list[1]:
        min1, min2 = mean_list[0], mean_list[1]
    else:
        min1, min2 = mean_list[1], mean_list[0]

    for x in mean_list[2:]:
        if x < min1:
            min2 = min1
            min1 = x
        elif x < min2:
            min2 = x
            
    return mean_list.index(min1), mean_list.index(min2)

def rank_sums_test(alg_data1, alg_data2):
    """执行Wilcoxon秩和检验
    
    Args:
        alg_data1: 第一组数据
        alg_data2: 第二组数据
        
    Returns:
        bool: 是否有显著差异
    """
    stat, p_value = ranksums(alg_data1, alg_data2)
    alpha = 0.05
    return p_value < alpha

def run(config):
    """运行MIGD指标计算
    
    Args:
        config: 配置信息，包含input_paths, output_path, n, tau等
    """
    input_paths = config["input_paths"]
    output_path = config["output_path"]
    your_algorithm = config.get("your_algorithm", "")  # 获取主体算法路径
    
    # 使用嵌套字典存储结果
    # 结构: {响应策略: {搜索算法: {问题名: {(n,tau): [MIGD值列表]}}}}
    results = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
    
    # 使用生成器处理文件加载
    for result in load_result_from_files(input_paths):
        # 从settings中获取信息
        settings = result.get('settings', {})
        response, search, problem, n, tau = extract_info_from_settings(settings)
        
        # 计算MIGD值
        migd_value = calculate_MIGD(result['runtime_populations'])
        
        # 存储结果
        results[response][search][problem][(n, tau)].append(migd_value)
    
    # 创建excel文件
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    
    # 获取所有唯一的算法组合
    algorithms = []
    for response in results:
        for search in results[response]:
            algorithms.append(f"{response}-{search}")
    
    # 如果指定了主体算法，将其放在第一位
    if your_algorithm:
        if your_algorithm in algorithms:
            algorithms.remove(your_algorithm)
            algorithms.insert(0, your_algorithm)
    
    # 创建表头
    worksheet.cell(1, 2, '(n,tau)')
    for i in range(3, 3 + len(algorithms)):
        worksheet.cell(1, i, algorithms[i - 3])
    
    # 获取所有问题名
    problem_names = set()
    for response in results:
        for search in results[response]:
            problem_names.update(results[response][search].keys())
    problem_names = sorted(list(problem_names))
    
    # 获取所有环境组合
    environments = set()
    for response in results:
        for search in results[response]:
            for problem in results[response][search]:
                environments.update(results[response][search][problem].keys())
    environments = sorted(list(environments))
    
    now_row = 2
    # rank_sums_test的记录
    rank_sums_collection = [{"win": 0, "loss": 0, "eq": 0} for _ in range(len(algorithms) - 1)]
    
    # 处理每个问题
    for problem in problem_names:
        now_column = 1
        # 合并单元格写问题名
        block = worksheet.cell(now_row, 1)
        worksheet.merge_cells(start_row=now_row, start_column=1, 
                            end_row=now_row + len(environments) - 1, end_column=1)
        block.value = problem
        block.alignment = Alignment(horizontal='center', vertical='center')
        
        begin_col = 2
        # 处理每个环境
        for env in environments:
            now_column = begin_col
            # 写环境名
            block = worksheet.cell(now_row, now_column)
            block.value = str(env)
            now_column += 1
            
            # 收集该环境下的所有算法结果
            alg_grades = []
            target_alg_grades = None
            target_alg_mean = None
            
            for i, alg in enumerate(algorithms):
                response, search = alg.split('-')
                migd_values = results[response][search][problem][env]
                mean = sum(migd_values) / len(migd_values)
                var = sum((x - mean) ** 2 for x in migd_values) / len(migd_values)
                alg_grades.append(mean)
                
                compare_mark = ''
                if i == 0:
                    target_alg_grades = migd_values
                    target_alg_mean = mean
                else:
                    bo = rank_sums_test(target_alg_grades, migd_values)
                    if bo:
                        if target_alg_mean > mean:
                            rank_sums_collection[i - 1]['loss'] += 1
                            compare_mark = '-'
                        else:
                            rank_sums_collection[i - 1]['win'] += 1
                            compare_mark = '+'
                    else:
                        rank_sums_collection[i - 1]['eq'] += 1
                        compare_mark = '~'
                
                mean_with_var = f"{mean:.3e}({var:.3e}){compare_mark}"
                block = worksheet.cell(now_row, now_column)
                block.value = mean_with_var
                now_column += 1
            
            # 添加颜色标记
            min1_index, min2_index = find_first_and_second(alg_grades)
            fill = PatternFill(start_color="909090", end_color="909090", fill_type="solid")
            worksheet.cell(now_row, 3 + min1_index).fill = fill
            fill = PatternFill(start_color="cbcbcb", end_color="cbcbcb", fill_type="solid")
            worksheet.cell(now_row, 3 + min2_index).fill = fill
            now_row += 1
    
    # 写入rank-sum统计
    block = worksheet.cell(now_row, 1)
    block.value = '+/~/-'
    block.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells(start_row=now_row, start_column=1, end_row=now_row, end_column=2)
    
    for i in range(len(algorithms)):
        if i == 0:
            worksheet.cell(now_row, 3 + i).value = "-"
        else:
            worksheet.cell(now_row, 3 + i).value = f"{rank_sums_collection[i-1]['win']}/{rank_sums_collection[i-1]['eq']}/{rank_sums_collection[i-1]['loss']}"
    
    # 保存excel文件
    output_file = f'{output_path}/MIGD指标对比_{int(time.time())}.xlsx'
    workbook.save(output_file)
    print(f"Results saved to: {output_file}")

